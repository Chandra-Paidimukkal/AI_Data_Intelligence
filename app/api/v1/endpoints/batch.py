"""
batch.py — Batch PDF extraction endpoint.

POST /api/v1/batch/run
  Upload multiple PDFs + schema_id → parse + extract all → returns batch_id.

POST /api/v1/batch/run-from-documents
  Run batch extraction on already-uploaded document IDs (no re-upload needed).

GET  /api/v1/batch/{batch_id}
  Poll batch status: { status, total, completed, failed, job_ids }

GET  /api/v1/batch/{batch_id}/excel
  Download combined Excel with ALL records from ALL documents.

GET  /api/v1/batch/{batch_id}/csv
  Download combined CSV.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import os
import shutil
import uuid
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.database import get_db
from app.models.document import Document
from app.models.job import ExtractionBatch, ExtractionJob
from app.models.schema import SchemaDefinition
from app.services.parser import parse_document
from app.services.pipeline import run_extraction

router = APIRouter(prefix="/batch", tags=["Batch"])


# ── Run batch ─────────────────────────────────────────────────────────────────

@router.post("/run", tags=["Documents"])
async def run_batch(
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(..., description="One or more PDF files"),
    schema_id: str = Form(..., description="ID of a saved schema"),
    provider: str = Form(default="landingai"),
    api_key: str = Form(default=""),
    base_url: str = Form(default=""),
    model: str = Form(default=""),
    multi_record: bool = Form(default=True, description="Extract all model variants per PDF"),
    db: Session = Depends(get_db),
):
    """
    Upload multiple PDFs and run extraction on all of them.
    Returns a batch_id immediately; processing runs in the background.
    Poll GET /batch/{batch_id} for status.
    """
    # Validate schema
    schema_def = db.query(SchemaDefinition).filter(SchemaDefinition.id == schema_id).first()
    if not schema_def:
        raise HTTPException(404, f"Schema '{schema_id}' not found.")
    schema_dict = schema_def.raw_definition or {"name": schema_def.name, "fields": schema_def.fields}

    if not files:
        raise HTTPException(400, "No files uploaded.")

    # Save all uploaded files and create Document records
    doc_ids = []
    for file in files:
        doc_id = str(uuid.uuid4())
        ext = Path(file.filename or "file.bin").suffix or ".pdf"
        fpath = os.path.join(settings.UPLOAD_DIR, f"{doc_id}{ext}")
        with open(fpath, "wb") as f:
            shutil.copyfileobj(file.file, f)

        doc = Document(
            id=doc_id,
            file_name=file.filename,
            file_path=fpath,
            file_size=os.path.getsize(fpath),
            mime_type=file.content_type or "application/pdf",
            status="uploaded",
        )
        db.add(doc)
        doc_ids.append((doc_id, fpath, file.filename))

    # Create batch record
    batch_id = str(uuid.uuid4())
    batch = ExtractionBatch(
        id=batch_id,
        schema_id=schema_id,
        document_ids=[d[0] for d in doc_ids],
        job_ids=[],
        status="running",
        total=len(doc_ids),
        completed=0,
        failed=0,
    )
    db.add(batch)
    db.commit()

    # Run processing in background
    provider_config = {
        "provider": provider,
        "api_key": api_key,
        "base_url": base_url,
        "model": model,
    }
    background_tasks.add_task(
        _process_batch,
        batch_id=batch_id,
        doc_ids=doc_ids,
        schema_dict=schema_dict,
        schema_id=schema_id,
        provider_config=provider_config,
        multi_record=multi_record,
    )

    return {
        "batch_id": batch_id,
        "total_files": len(doc_ids),
        "status": "running",
        "message": f"Processing {len(doc_ids)} file(s) in background.",
        "poll_url": f"/api/v1/batch/{batch_id}",
        "export_url": f"/api/v1/batch/{batch_id}/excel",
    }


# ── Run batch from existing document IDs ─────────────────────────────────────

class BatchFromDocumentsRequest(BaseModel):
    document_ids: list[str]
    schema_id: str
    provider: str = "landingai"
    api_key: str = ""
    base_url: str = ""
    model: str = ""
    multi_record: bool = True


@router.post("/run-from-documents", tags=["Extraction"])
async def run_batch_from_documents(
    req: BatchFromDocumentsRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    """
    Run batch extraction on already-uploaded/parsed documents.
    Pass a list of document_ids instead of re-uploading files.
    """
    if not req.document_ids:
        raise HTTPException(400, "No document_ids provided.")

    # Validate schema
    schema_def = db.query(SchemaDefinition).filter(SchemaDefinition.id == req.schema_id).first()
    if not schema_def:
        raise HTTPException(404, f"Schema '{req.schema_id}' not found.")
    schema_dict = schema_def.raw_definition or {"name": schema_def.name, "fields": schema_def.fields}

    # Validate documents exist
    doc_ids = []
    for doc_id in req.document_ids:
        doc = db.query(Document).filter(Document.id == doc_id).first()
        if not doc:
            raise HTTPException(404, f"Document '{doc_id}' not found.")
        doc_ids.append((doc_id, doc.file_path, doc.file_name))

    # Create batch
    batch_id = str(uuid.uuid4())
    batch = ExtractionBatch(
        id=batch_id,
        schema_id=req.schema_id,
        document_ids=req.document_ids,
        job_ids=[],
        status="running",
        total=len(doc_ids),
        completed=0,
        failed=0,
    )
    db.add(batch)
    db.commit()

    provider_config = {
        "provider": req.provider,
        "api_key": req.api_key,
        "base_url": req.base_url,
        "model": req.model,
    }
    background_tasks.add_task(
        _process_batch,
        batch_id=batch_id,
        doc_ids=doc_ids,
        schema_dict=schema_dict,
        schema_id=req.schema_id,
        provider_config=provider_config,
        multi_record=req.multi_record,
    )

    return {
        "batch_id": batch_id,
        "total_files": len(doc_ids),
        "status": "running",
        "message": f"Processing {len(doc_ids)} document(s) in background.",
        "poll_url": f"/api/v1/batch/{batch_id}",
        "export_url": f"/api/v1/batch/{batch_id}/excel",
    }




async def _process_batch(
    batch_id: str,
    doc_ids: list,
    schema_dict: dict,
    schema_id: str,
    provider_config: dict,
    multi_record: bool,
):
    """Process all documents in the batch sequentially."""
    from app.core.database import SessionLocal
    db = SessionLocal()

    try:
        job_ids = []
        completed = 0
        failed = 0

        for doc_id, fpath, filename in doc_ids:
            try:
                # Parse document (skip if already parsed)
                doc = db.query(Document).filter(Document.id == doc_id).first()
                if doc and doc.status == "parsed" and doc.parsed_data:
                    parsed = doc.parsed_data
                else:
                    parsed = parse_document(fpath)
                    if doc:
                        doc.status = "parsed"
                        doc.parsed_data = parsed
                        doc.page_count = parsed["metadata"].get("page_count", 1)
                        db.commit()

                # Create job
                job_id = str(uuid.uuid4())
                job = ExtractionJob(
                    id=job_id,
                    document_id=doc_id,
                    schema_name=schema_dict.get("name", "inline"),
                    schema_id=schema_id,
                    batch_id=batch_id,
                    status="running",
                    provider=provider_config.get("provider", "none"),
                )
                db.add(job)
                db.commit()

                # Run extraction
                provider = provider_config.get("provider", "none").lower()

                if provider == "landingai":
                    from app.services.landingai_service import (
                        extract_with_landingai, extract_multi_with_landingai
                    )
                    api_key = provider_config.get("api_key", "")
                    environment = provider_config.get("base_url", "production")
                    markdown = parsed.get("markdown", "")

                    if multi_record:
                        extraction = await extract_multi_with_landingai(
                            markdown=markdown,
                            schema=schema_dict,
                            api_key=api_key,
                            environment=environment,
                        )
                    else:
                        extraction = await extract_with_landingai(
                            markdown=markdown,
                            schema=schema_dict,
                            api_key=api_key,
                            environment=environment,
                        )
                else:
                    extraction = await run_extraction(
                        parsed_doc=parsed,
                        schema=schema_dict,
                        provider_config=provider_config,
                    )

                # Attach source filename to each record for traceability
                if "records" in extraction:
                    for rec in extraction["records"]:
                        rec["source_file"] = filename
                else:
                    extraction["source_file"] = filename

                job.status = "completed"
                job.result = extraction
                db.commit()

                job_ids.append(job_id)
                completed += 1

            except Exception as e:
                failed += 1
                # Try to mark job as failed if it was created
                try:
                    job.status = "failed"
                    job.error = str(e)
                    db.commit()
                    job_ids.append(job_id)
                except Exception:
                    pass

            # Update batch progress
            batch = db.query(ExtractionBatch).filter(ExtractionBatch.id == batch_id).first()
            if batch:
                batch.completed = completed
                batch.failed = failed
                batch.job_ids = job_ids
                db.commit()

        # Mark batch complete
        batch = db.query(ExtractionBatch).filter(ExtractionBatch.id == batch_id).first()
        if batch:
            batch.status = "completed" if failed == 0 else "completed_with_errors"
            batch.job_ids = job_ids
            db.commit()

    except Exception as e:
        batch = db.query(ExtractionBatch).filter(ExtractionBatch.id == batch_id).first()
        if batch:
            batch.status = "failed"
            db.commit()
    finally:
        db.close()


# ── List all batches ──────────────────────────────────────────────────────────

@router.get("", tags=["Extraction"])
async def list_batches(db: Session = Depends(get_db)):
    """List all batch sessions with status and job count."""
    from app.models.job import ExtractionBatch
    batches = db.query(ExtractionBatch).order_by(ExtractionBatch.created_at.desc()).limit(100).all()
    return {"batches": [
        {
            "batch_id": b.id,
            "schema_id": b.schema_id,
            "status": b.status,
            "total": b.total,
            "completed": b.completed,
            "failed": b.failed,
            "job_ids": b.job_ids or [],
            "created_at": b.created_at.isoformat() if b.created_at else None,
        }
        for b in batches
    ]}


# ── Upload ZIP and run batch ──────────────────────────────────────────────────

@router.post("/run-from-zip", tags=["Documents"])
async def run_batch_from_zip(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="ZIP file containing PDFs"),
    schema_id: str = Form(..., description="ID of a saved schema"),
    provider: str = Form(default="landingai"),
    api_key: str = Form(default=""),
    base_url: str = Form(default=""),
    model: str = Form(default=""),
    multi_record: bool = Form(default=True),
    db: Session = Depends(get_db),
):
    """
    Upload a ZIP file containing PDFs.
    Extracts all PDFs from the ZIP, parses them, and runs batch extraction.
    """
    import zipfile

    if not file.filename or not file.filename.lower().endswith(".zip"):
        raise HTTPException(400, "File must be a .zip archive.")

    schema_def = db.query(SchemaDefinition).filter(SchemaDefinition.id == schema_id).first()
    if not schema_def:
        raise HTTPException(404, f"Schema '{schema_id}' not found.")
    schema_dict = schema_def.raw_definition or {"name": schema_def.name, "fields": schema_def.fields}

    # Save the ZIP temporarily
    zip_id = str(uuid.uuid4())
    zip_path = os.path.join(settings.UPLOAD_DIR, f"{zip_id}.zip")
    with open(zip_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # Extract PDFs from ZIP
    doc_ids = []
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            pdf_names = [
                name for name in zf.namelist()
                if not name.startswith("__MACOSX") and not name.startswith(".")
                and name.lower().endswith((".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".docx", ".xlsx"))
            ]
            if not pdf_names:
                raise HTTPException(400, "ZIP contains no supported files (PDF, PNG, JPG, DOCX, XLSX).")

            for name in pdf_names:
                doc_id = str(uuid.uuid4())
                ext = Path(name).suffix or ".pdf"
                fpath = os.path.join(settings.UPLOAD_DIR, f"{doc_id}{ext}")
                with zf.open(name) as src, open(fpath, "wb") as dst:
                    shutil.copyfileobj(src, dst)

                basename = Path(name).name
                doc = Document(
                    id=doc_id,
                    file_name=basename,
                    file_path=fpath,
                    file_size=os.path.getsize(fpath),
                    mime_type="application/pdf",
                    status="uploaded",
                )
                db.add(doc)
                doc_ids.append((doc_id, fpath, basename))
    except zipfile.BadZipFile:
        raise HTTPException(400, "Invalid ZIP file.")
    finally:
        # Clean up the ZIP
        try:
            os.remove(zip_path)
        except Exception:
            pass

    if not doc_ids:
        raise HTTPException(400, "No valid files found in ZIP.")

    db.commit()

    # Create batch
    batch_id = str(uuid.uuid4())
    batch = ExtractionBatch(
        id=batch_id,
        schema_id=schema_id,
        document_ids=[d[0] for d in doc_ids],
        job_ids=[],
        status="running",
        total=len(doc_ids),
        completed=0,
        failed=0,
    )
    db.add(batch)
    db.commit()

    provider_config = {
        "provider": provider, "api_key": api_key,
        "base_url": base_url, "model": model,
    }
    background_tasks.add_task(
        _process_batch,
        batch_id=batch_id,
        doc_ids=doc_ids,
        schema_dict=schema_dict,
        schema_id=schema_id,
        provider_config=provider_config,
        multi_record=multi_record,
    )

    return {
        "batch_id": batch_id,
        "total_files": len(doc_ids),
        "status": "running",
        "message": f"Extracted {len(doc_ids)} file(s) from ZIP. Processing in background.",
        "poll_url": f"/api/v1/batch/{batch_id}",
        "export_url": f"/api/v1/batch/{batch_id}/excel",
    }


# ── Status ────────────────────────────────────────────────────────────────────

@router.get("/{batch_id}", tags=["Extraction"])
async def get_batch_status(batch_id: str, db: Session = Depends(get_db)):
    """Poll batch processing status."""
    batch = db.query(ExtractionBatch).filter(ExtractionBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(404, f"Batch '{batch_id}' not found.")

    return {
        "batch_id": batch_id,
        "status": batch.status,
        "total": batch.total,
        "completed": batch.completed,
        "failed": batch.failed,
        "job_ids": batch.job_ids or [],
        "schema_id": batch.schema_id,
        "created_at": batch.created_at.isoformat() if batch.created_at else None,
        "export_excel": f"/api/v1/batch/{batch_id}/excel" if batch.status in ("completed", "completed_with_errors") else None,
        "export_csv": f"/api/v1/batch/{batch_id}/csv" if batch.status in ("completed", "completed_with_errors") else None,
    }


@router.get("/{batch_id}/wait", tags=["Extraction"])
async def wait_for_batch(batch_id: str, db: Session = Depends(get_db)):
    """
    Wait until the batch is fully complete, then return the final status.
    Polls internally every 3 seconds for up to 10 minutes.
    Use this instead of manually polling GET /{batch_id}.
    """
    import asyncio
    max_wait = 600  # 10 minutes
    interval = 3
    elapsed = 0

    while elapsed < max_wait:
        # Re-query each iteration to get fresh data
        db.expire_all()
        batch = db.query(ExtractionBatch).filter(ExtractionBatch.id == batch_id).first()
        if not batch:
            raise HTTPException(404, f"Batch '{batch_id}' not found.")

        if batch.status in ("completed", "completed_with_errors", "failed"):
            return {
                "batch_id": batch_id,
                "status": batch.status,
                "total": batch.total,
                "completed": batch.completed,
                "failed": batch.failed,
                "job_ids": batch.job_ids or [],
                "schema_id": batch.schema_id,
                "created_at": batch.created_at.isoformat() if batch.created_at else None,
                "export_excel": f"/api/v1/batch/{batch_id}/excel" if batch.status in ("completed", "completed_with_errors") else None,
                "export_csv": f"/api/v1/batch/{batch_id}/csv" if batch.status in ("completed", "completed_with_errors") else None,
            }

        await asyncio.sleep(interval)
        elapsed += interval

    raise HTTPException(408, f"Batch did not complete within {max_wait} seconds.")


# ── Combined export helpers ───────────────────────────────────────────────────

def _collect_all_rows(batch: ExtractionBatch, db: Session) -> tuple[list[dict], list[str]]:
    """
    Collect all rows from all completed jobs in the batch.
    Returns (rows, field_names).
    Each row includes a 'source_file' column for traceability.
    Handles nested model arrays (e.g. result.models, result.items, result.products).
    """
    field_names: list[str] = []
    all_rows: list[dict] = []

    for job_id in (batch.job_ids or []):
        job = db.query(ExtractionJob).filter(ExtractionJob.id == job_id).first()
        if not job or job.status != "completed" or not job.result:
            continue

        result = job.result
        source_file = result.get("source_file", "")

        # Get schema field names from first job
        if not field_names:
            field_names = result.get("schema_fields", [])

        # Multi-record result: { records: [{result, confidence}, ...] }
        if "records" in result:
            for rec in result["records"]:
                rec_result = rec.get("result", {})
                rec_source = rec.get("source_file", source_file)
                row = {"source_file": rec_source}
                row.update({f: rec_result.get(f) for f in field_names})
                all_rows.append(row)

        # Single-record result: { result: {...} }
        elif "result" in result:
            rec_result = result.get("result", {})

            # Check for nested array of models (e.g. rec_result["models"] = [{...}, ...])
            array_key = None
            array_val = None
            for k, v in rec_result.items():
                if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                    array_key = k
                    array_val = v
                    break

            if array_key and array_val:
                # Top-level fields repeated on every row
                top_fields = {k: v for k, v in rec_result.items() if k != array_key}
                # Collect all item-level keys
                item_keys: list[str] = []
                for item in array_val:
                    for k in item.keys():
                        if k not in item_keys:
                            item_keys.append(k)
                # Update field_names if not yet set
                if not field_names:
                    field_names = list(top_fields.keys()) + item_keys

                for item in array_val:
                    row = {"source_file": source_file}
                    row.update({k: top_fields.get(k) for k in top_fields})
                    row.update({k: item.get(k) for k in item_keys})
                    all_rows.append(row)
            else:
                row = {"source_file": source_file}
                row.update({f: rec_result.get(f) for f in field_names})
                all_rows.append(row)

    # Build final column list: source_file first, then all field names
    # Collect any extra keys found in rows that aren't in field_names
    extra_keys: list[str] = []
    for row in all_rows:
        for k in row:
            if k != "source_file" and k not in field_names and k not in extra_keys:
                extra_keys.append(k)

    all_columns = ["source_file"] + field_names + extra_keys
    return all_rows, all_columns


# ── Excel export ──────────────────────────────────────────────────────────────

@router.get("/{batch_id}/excel", tags=["Export"])
async def export_batch_excel(batch_id: str, db: Session = Depends(get_db)):
    """
    Download combined Excel for all documents in the batch.
    One row per model/record, schema fields as column headers.
    """
    batch = db.query(ExtractionBatch).filter(ExtractionBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(404, f"Batch '{batch_id}' not found.")
    if batch.status not in ("completed", "completed_with_errors"):
        raise HTTPException(400, f"Batch not ready. Status: {batch.status}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed.")

    rows, columns = _collect_all_rows(batch, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Results"

    # Styles
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    row_fill_even = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(horizontal="left", vertical="center")

    # Headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        for col_idx, col_name in enumerate(columns, start=1):
            val = row.get(col_name)
            if val is None:
                val = ""
            elif isinstance(val, (list, dict)):
                val = json.dumps(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = data_align
            cell.border = thin

    # Auto-size columns
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row in rows:
            val = row.get(col_name)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=batch_{batch_id}.xlsx"},
    )


# ── CSV export ────────────────────────────────────────────────────────────────

@router.get("/{batch_id}/csv", tags=["Export"])
async def export_batch_csv(batch_id: str, db: Session = Depends(get_db)):
    """Download combined CSV for all documents in the batch."""
    batch = db.query(ExtractionBatch).filter(ExtractionBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(404, f"Batch '{batch_id}' not found.")
    if batch.status not in ("completed", "completed_with_errors"):
        raise HTTPException(400, f"Batch not ready. Status: {batch.status}")

    rows, columns = _collect_all_rows(batch, db)

    import csv
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: "" if row.get(c) is None else row.get(c) for c in columns})

    content = output.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=batch_{batch_id}.csv"},
    )

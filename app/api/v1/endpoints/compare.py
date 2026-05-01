"""
compare.py — Multi-engine extraction comparison.

POST /api/v1/compare/run
  Run extraction on the same document using multiple engines simultaneously.
  Returns side-by-side results for each engine so you can compare accuracy.

GET /api/v1/compare/{compare_id}
  Retrieve a previous comparison result.

GET /api/v1/compare/{compare_id}/excel
  Download comparison as Excel — one sheet per engine + a summary sheet.
"""
from __future__ import annotations

import asyncio
import io
import json
import time
import uuid
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.document import Document
from app.models.job import ExtractionJob
from app.models.schema import SchemaDefinition
from app.services.pipeline import run_extraction

router = APIRouter(prefix="/compare", tags=["Compare"])


# ── Request models ────────────────────────────────────────────────────────────

class EngineConfig(BaseModel):
    provider: str = Field(..., description="Engine: openai | anthropic | gemini | groq | grok | emergence | landingai | python | hybrid")
    api_key: str = Field(default="")
    model: str = Field(default="")
    base_url: str = Field(default="")
    label: str = Field(default="", description="Optional display name for this engine")


class CompareRequest(BaseModel):
    document_id: str
    schema_id: Optional[str] = None
    schema: Optional[dict] = None
    engines: list[EngineConfig] = Field(..., description="List of engines to compare")
    options: dict = Field(default_factory=dict)


# ── In-memory store for comparison results (use DB job for persistence) ───────
_compare_store: dict[str, dict] = {}


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/run")
async def run_comparison(
    req: CompareRequest,
    db: Session = Depends(get_db),
):
    """
    Run extraction using multiple engines on the same document.
    All engines run in parallel. Returns compare_id + all results.
    """
    # Resolve document
    doc = db.query(Document).filter(Document.id == req.document_id).first()
    if not doc:
        raise HTTPException(404, f"Document '{req.document_id}' not found.")
    if doc.status != "parsed" or not doc.parsed_data:
        raise HTTPException(400, f"Document not parsed. Status: {doc.status}")

    # Resolve schema
    if req.schema_id:
        schema_def = db.query(SchemaDefinition).filter(SchemaDefinition.id == req.schema_id).first()
        if not schema_def:
            raise HTTPException(404, f"Schema '{req.schema_id}' not found.")
        schema_dict = schema_def.raw_definition or {"name": schema_def.name, "fields": schema_def.fields}
    elif req.schema:
        schema_dict = req.schema
    else:
        raise HTTPException(400, "Provide either 'schema_id' or 'schema'.")

    if not req.engines:
        raise HTTPException(400, "Provide at least one engine in 'engines'.")

    compare_id = str(uuid.uuid4())
    start = time.time()

    # Run all engines in parallel
    tasks = [
        _run_engine(engine, doc.parsed_data, schema_dict, req.options)
        for engine in req.engines
    ]
    engine_results = await asyncio.gather(*tasks, return_exceptions=False)

    # Build comparison response
    results = []
    for engine, result in zip(req.engines, engine_results):
        label = engine.label or engine.provider
        results.append({
            "engine": engine.provider,
            "label": label,
            "model": engine.model or _default_model(engine.provider),
            "status": result.get("status", "completed"),
            "error": result.get("error"),
            "result": result.get("result", {}),
            "confidence": result.get("confidence", {}),
            "sources": result.get("sources", {}),
            "duration_seconds": result.get("duration_seconds", 0),
            "schema_fields": result.get("schema_fields", []),
        })

    schema_fields = schema_dict.get("fields", [])
    field_names = [f["name"] for f in schema_fields]

    comparison = {
        "compare_id": compare_id,
        "document_id": req.document_id,
        "schema_name": schema_dict.get("name", "inline"),
        "total_engines": len(results),
        "duration_seconds": round(time.time() - start, 2),
        "field_names": field_names,
        "engines": results,
        "summary": _build_summary(results, field_names),
    }

    # Store for later retrieval
    _compare_store[compare_id] = comparison

    return comparison


@router.get("/{compare_id}")
async def get_comparison(compare_id: str):
    """Retrieve a previous comparison result."""
    result = _compare_store.get(compare_id)
    if not result:
        raise HTTPException(404, f"Comparison '{compare_id}' not found.")
    return result


@router.get("/{compare_id}/excel")
async def export_comparison_excel(compare_id: str):
    """
    Download comparison as Excel.
    - Sheet 1: Summary — all engines side by side, one row per field
    - Sheet 2+: One sheet per engine with full results
    """
    comparison = _compare_store.get(compare_id)
    if not comparison:
        raise HTTPException(404, f"Comparison '{compare_id}' not found.")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed.")

    wb = openpyxl.Workbook()

    # Styles
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    data_font = Font(name="Calibri", size=10)
    row_even = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")
    row_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    field_names = comparison["field_names"]
    engines = comparison["engines"]

    # ── Sheet 1: Summary (field × engine matrix) ──────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    # Headers: Field | Engine1 | Engine2 | ...
    headers = ["Field"] + [e["label"] for e in engines]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    # One row per field
    for row_idx, fname in enumerate(field_names, 2):
        fill = row_even if row_idx % 2 == 0 else row_odd

        # Field name
        cell = ws.cell(row=row_idx, column=1, value=fname)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = fill
        cell.border = thin

        # Value from each engine
        for col_idx, engine in enumerate(engines, 2):
            val = engine["result"].get(fname)
            conf = engine["confidence"].get(fname, 0)
            display = "" if val is None else val
            if isinstance(display, (list, dict)):
                display = json.dumps(display)

            cell = ws.cell(row=row_idx, column=col_idx, value=display)
            cell.font = data_font
            cell.border = thin
            cell.alignment = Alignment(wrap_text=False)

            # Color by confidence
            if val is None:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif conf >= 0.8:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif conf >= 0.5:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = fill

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30

    ws.freeze_panes = "B2"

    # ── Sheet per engine ──────────────────────────────────────────────────────
    for engine in engines:
        ws_e = wb.create_sheet(title=engine["label"][:31])  # Excel sheet name limit

        # Header row
        eng_headers = ["Field", "Value", "Confidence", "Source"]
        for col, h in enumerate(eng_headers, 1):
            cell = ws_e.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin

        # Engine info row
        ws_e.cell(row=2, column=1, value="Engine").font = Font(bold=True)
        ws_e.cell(row=2, column=2, value=f"{engine['provider']} / {engine['model']}")
        ws_e.cell(row=3, column=1, value="Duration").font = Font(bold=True)
        ws_e.cell(row=3, column=2, value=f"{engine['duration_seconds']}s")
        if engine.get("error"):
            ws_e.cell(row=4, column=1, value="Error").font = Font(bold=True, color="FF0000")
            ws_e.cell(row=4, column=2, value=engine["error"])

        start_row = 6
        for row_idx, fname in enumerate(field_names, start_row):
            fill = row_even if row_idx % 2 == 0 else row_odd
            val = engine["result"].get(fname)
            conf = engine["confidence"].get(fname, 0)
            source = engine["sources"].get(fname, "")
            display = "" if val is None else val
            if isinstance(display, (list, dict)):
                display = json.dumps(display)

            ws_e.cell(row=row_idx, column=1, value=fname).font = Font(bold=True, size=10)
            ws_e.cell(row=row_idx, column=2, value=display).font = data_font
            ws_e.cell(row=row_idx, column=3, value=conf).font = data_font
            ws_e.cell(row=row_idx, column=4, value=source).font = data_font

            for col in range(1, 5):
                ws_e.cell(row=row_idx, column=col).fill = fill
                ws_e.cell(row=row_idx, column=col).border = thin

        ws_e.column_dimensions["A"].width = 35
        ws_e.column_dimensions["B"].width = 45
        ws_e.column_dimensions["C"].width = 12
        ws_e.column_dimensions["D"].width = 20
        ws_e.freeze_panes = "A6"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=comparison_{compare_id}.xlsx"},
    )


# ── Engine runner ─────────────────────────────────────────────────────────────

async def _run_engine(
    engine: EngineConfig,
    parsed_doc: dict,
    schema_dict: dict,
    options: dict,
) -> dict:
    """Run a single engine and return standardized result."""
    start = time.time()
    provider = engine.provider.lower()

    try:
        if provider == "landingai":
            from app.services.landingai_service import (
                extract_with_landingai, extract_multi_with_landingai
            )
            markdown = parsed_doc.get("markdown", "")
            environment = engine.base_url or "production"
            multi = options.get("multi_record", False)

            if multi:
                raw = await extract_multi_with_landingai(
                    markdown=markdown, schema=schema_dict,
                    api_key=engine.api_key, environment=environment,
                )
            else:
                raw = await extract_with_landingai(
                    markdown=markdown, schema=schema_dict,
                    api_key=engine.api_key, environment=environment,
                )
            return {**raw, "status": "completed", "duration_seconds": round(time.time() - start, 2)}

        else:
            provider_config = {
                "provider": provider,
                "api_key": engine.api_key,
                "model": engine.model,
                "base_url": engine.base_url,
            }
            raw = await run_extraction(
                parsed_doc=parsed_doc,
                schema=schema_dict,
                provider_config=provider_config,
            )
            return {**raw, "status": "completed"}

    except Exception as e:
        fields = schema_dict.get("fields", [])
        field_names = [f["name"] for f in fields]
        return {
            "status": "failed",
            "error": str(e),
            "result": {f: None for f in field_names},
            "confidence": {f: 0.0 for f in field_names},
            "sources": {f: provider for f in field_names},
            "schema_fields": field_names,
            "duration_seconds": round(time.time() - start, 2),
        }


# ── Summary builder ───────────────────────────────────────────────────────────

def _build_summary(results: list[dict], field_names: list[str]) -> dict:
    """
    Build a summary showing:
    - Which fields were extracted by how many engines
    - Agreement score per field (% of engines that got the same value)
    - Best engine overall (most non-null fields)
    """
    field_coverage: dict[str, int] = {}
    field_agreement: dict[str, float] = {}
    engine_scores: dict[str, int] = {}

    for fname in field_names:
        values = [r["result"].get(fname) for r in results if r.get("result")]
        non_null = [v for v in values if v is not None]
        field_coverage[fname] = len(non_null)

        # Agreement: % of non-null values that match the most common value
        if non_null:
            from collections import Counter
            most_common_count = Counter(str(v) for v in non_null).most_common(1)[0][1]
            field_agreement[fname] = round(most_common_count / len(non_null), 2)
        else:
            field_agreement[fname] = 0.0

    for r in results:
        label = r.get("label", r.get("engine", "unknown"))
        non_null = sum(1 for v in r.get("result", {}).values() if v is not None)
        engine_scores[label] = non_null

    best_engine = max(engine_scores, key=engine_scores.get) if engine_scores else None

    return {
        "field_coverage": field_coverage,
        "field_agreement": field_agreement,
        "engine_scores": engine_scores,
        "best_engine": best_engine,
        "total_fields": len(field_names),
    }


def _default_model(provider: str) -> str:
    defaults = {
        "openai": "gpt-4o-mini", "chatgpt": "gpt-4o-mini",
        "anthropic": "claude-3-5-haiku-20241022",
        "gemini": "gemini-1.5-flash",
        "groq": "llama-3.1-8b-instant",
        "grok": "grok-beta",
        "emergence": "em-llm-001",
        "landingai": "dpt-2-latest",
        "python": "heuristic",
        "hybrid": "heuristic+ai",
    }
    return defaults.get(provider, provider)

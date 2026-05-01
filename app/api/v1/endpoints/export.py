"""
export.py — CSV and Excel export for extraction jobs.

Supports both single-record and multi-record (multi_record: true) jobs.
Each schema field becomes a column header; each record becomes a row.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import json
import csv
import io
from typing import Any

from app.core.database import get_db
from app.models.job import ExtractionJob

router = APIRouter(prefix="/export", tags=["Export"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _extract_rows(job_result: dict) -> tuple[list[dict], list[str]]:
    """
    Extract flat rows and ordered field names from a job result.
    Handles both single-record and multi-record results.
    Also handles schemas where models/items are nested under any array key
    (e.g. 'models', 'records', 'items', 'products').

    Returns (rows, field_names) where each row is a flat dict of field→value.
    """
    if not job_result:
        return [], []

    # Multi-record mode: { records: [{result, confidence, ...}, ...] }
    if "records" in job_result:
        records = job_result["records"]
        field_names = job_result.get("schema_fields", [])
        if not field_names and records:
            field_names = records[0].get("schema_fields", list(records[0].get("result", {}).keys()))

        rows = []
        for rec in records:
            result = rec.get("result", {})
            rows.append({f: result.get(f) for f in field_names})
        return rows, field_names

    # Single-record mode: { result: {...}, schema_fields: [...] }
    if "result" in job_result:
        result = job_result["result"]
        schema_fields = job_result.get("schema_fields", [])

        # Check if any field in the result is a list of dicts (e.g. "models", "items", "products")
        # If so, flatten: top-level fields + each item in the array as a row
        array_key = None
        array_val = None
        for k, v in result.items():
            if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                array_key = k
                array_val = v
                break

        if array_key and array_val:
            # Top-level fields (everything except the array)
            top_fields = {k: v for k, v in result.items() if k != array_key}
            # Get all keys from the array items
            item_keys = []
            for item in array_val:
                for k in item.keys():
                    if k not in item_keys:
                        item_keys.append(k)

            # Build field names: top-level first, then item fields
            top_field_names = list(top_fields.keys())
            field_names = top_field_names + item_keys

            # One row per model/item, with top-level fields repeated
            rows = []
            for item in array_val:
                row = {k: top_fields.get(k) for k in top_field_names}
                for k in item_keys:
                    row[k] = item.get(k)
                rows.append(row)
            return rows, field_names

        # Plain single record — no nested arrays
        field_names = schema_fields if schema_fields else list(result.keys())
        rows = [{f: result.get(f) for f in field_names}]
        return rows, field_names

    return [], []


def _clean_value(v: Any) -> str:
    """Convert a value to a clean string for CSV/Excel."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float)):
        return v  # keep numeric for Excel
    if isinstance(v, (list, dict)):
        return json.dumps(v)
    return str(v)


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/{job_id}/json")
async def export_json(job_id: str, db: Session = Depends(get_db)):
    """Export full job result as JSON."""
    job = db.query(ExtractionJob).filter(ExtractionJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")
    if job.status != "completed":
        raise HTTPException(400, f"Job not completed. Status: {job.status}")

    content = json.dumps(job.result or {}, indent=2, default=str)
    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=extraction_{job_id}.json"},
    )


@router.get("/{job_id}/csv")
async def export_csv(job_id: str, db: Session = Depends(get_db)):
    """
    Export extraction results as CSV.
    Each schema field is a column; each model/record is a row.
    """
    job = db.query(ExtractionJob).filter(ExtractionJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")
    if job.status != "completed":
        raise HTTPException(400, f"Job not completed. Status: {job.status}")

    rows, field_names = _extract_rows(job.result or {})
    if not rows:
        rows, field_names = [{}], []

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=field_names, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({f: "" if row.get(f) is None else row.get(f) for f in field_names})

    content = output.getvalue().encode("utf-8-sig")  # utf-8-sig for Excel compatibility
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=extraction_{job_id}.csv"},
    )


@router.get("/{job_id}/excel")
async def export_excel(job_id: str, db: Session = Depends(get_db)):
    """
    Export extraction results as Excel (.xlsx).
    Sheet 1 — Results: schema fields as headers, one row per model/record.
    Sheet 2 — Confidence: per-field confidence scores.
    """
    job = db.query(ExtractionJob).filter(ExtractionJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")
    if job.status != "completed":
        raise HTTPException(400, f"Job not completed. Status: {job.status}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    rows, field_names = _extract_rows(job.result or {})
    if not rows:
        rows, field_names = [{}], []

    wb = openpyxl.Workbook()

    # ── Sheet 1: Results ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Extraction Results"

    # Header style — dark blue background, white bold text
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Alternate row fills
    row_fill_even = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Write headers
    for col_idx, field in enumerate(field_names, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        for col_idx, field in enumerate(field_names, start=1):
            val = row.get(field)
            # Keep numbers as numbers, convert None to empty string
            if val is None:
                val = ""
            elif isinstance(val, (list, dict)):
                val = json.dumps(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = data_align
            cell.border = thin_border

    # Auto-size columns (cap at 40 chars wide)
    for col_idx, field in enumerate(field_names, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(field))
        for row in rows:
            val = row.get(field)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Confidence ───────────────────────────────────────────────────
    job_result = job.result or {}

    # For multi-record, show confidence per record
    if "records" in job_result:
        ws2 = wb.create_sheet("Confidence")
        conf_headers = ["Record #", "Model Number"] + field_names
        for col_idx, h in enumerate(conf_headers, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for rec_idx, rec in enumerate(job_result["records"], start=1):
            conf = rec.get("confidence", {})
            result = rec.get("result", {})
            model_num = result.get("model_number", f"Record {rec_idx}")
            row_data = [rec_idx, model_num] + [conf.get(f, "") for f in field_names]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws2.cell(row=rec_idx + 1, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                # Color-code confidence: green ≥0.8, yellow 0.5-0.8, red <0.5
                if isinstance(val, (int, float)) and col_idx > 2:
                    if val >= 0.8:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif val >= 0.5:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws2.freeze_panes = "A2"

    elif "confidence" in job_result:
        ws2 = wb.create_sheet("Confidence")
        ws2.append(["Field", "Confidence", "Source"])
        for field in field_names:
            conf = (job_result.get("confidence") or {}).get(field, "")
            source = (job_result.get("sources") or {}).get(field, "")
            ws2.append([field, conf, source])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=extraction_{job_id}.xlsx"},
    )

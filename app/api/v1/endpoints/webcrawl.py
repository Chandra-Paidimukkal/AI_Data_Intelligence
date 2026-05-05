"""
webcrawl.py — Web crawl extraction endpoint.

POST /api/v1/webcrawl/run
  Submit a web crawl job (URL list or deep crawl).

GET  /api/v1/webcrawl/{job_id}
  Poll job status.

GET  /api/v1/webcrawl/{job_id}/excel
GET  /api/v1/webcrawl/{job_id}/csv
  Download results.
"""
from __future__ import annotations

import csv
import io
import json
import uuid
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import Column, String, Integer, JSON, DateTime, Text
from sqlalchemy.orm import Session
from datetime import datetime

from app.core.database import get_db, Base
from app.core.auth import get_current_user_optional
from app.models.schema import SchemaDefinition
from app.models.user import User

router = APIRouter(prefix="/webcrawl", tags=["Web Crawl"])


# ── In-memory job store (simple, no DB model needed) ─────────────────────────
_jobs: dict[str, dict] = {}


# ── Request models ────────────────────────────────────────────────────────────

class WebCrawlRequest(BaseModel):
    urls: list[str]
    schema_id: str
    max_depth: int = 1  # 1 = single page, 2-5 = deep crawl
    max_pages: int = 50


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/run")
async def run_webcrawl(
    req: WebCrawlRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    """Start a web crawl extraction job."""
    schema = db.query(SchemaDefinition).filter(SchemaDefinition.id == req.schema_id).first()
    if not schema:
        raise HTTPException(404, f"Schema '{req.schema_id}' not found.")

    schema_dict = schema.raw_definition or {"name": schema.name, "fields": schema.fields}

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {
        "job_id": job_id,
        "status": "running",
        "total": len(req.urls),
        "completed": 0,
        "failed": 0,
        "results": [],
        "schema_name": schema.name,
        "created_at": datetime.utcnow().isoformat(),
        "user_id": current_user.id if current_user else None,
    }

    background_tasks.add_task(
        _process_crawl,
        job_id=job_id,
        urls=req.urls,
        schema_dict=schema_dict,
        max_depth=min(req.max_depth, 5),
        max_pages=min(req.max_pages, 200),
    )

    return {
        "job_id": job_id,
        "total_urls": len(req.urls),
        "status": "running",
        "poll_url": f"/api/v1/webcrawl/{job_id}",
    }


@router.post("/run-from-file")
async def run_webcrawl_from_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    schema_id: str = Form(...),
    max_depth: int = Form(default=1),
    max_pages: int = Form(default=50),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    """Upload a CSV/TXT/Excel file containing URLs and start crawl."""
    schema = db.query(SchemaDefinition).filter(SchemaDefinition.id == schema_id).first()
    if not schema:
        raise HTTPException(404, f"Schema '{schema_id}' not found.")

    schema_dict = schema.raw_definition or {"name": schema.name, "fields": schema.fields}

    content = await file.read()
    urls = _extract_urls_from_file(content, file.filename or "")

    if not urls:
        raise HTTPException(400, "No valid URLs found in the uploaded file.")

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {
        "job_id": job_id,
        "status": "running",
        "total": len(urls),
        "completed": 0,
        "failed": 0,
        "results": [],
        "schema_name": schema.name,
        "created_at": datetime.utcnow().isoformat(),
        "user_id": current_user.id if current_user else None,
    }

    background_tasks.add_task(
        _process_crawl,
        job_id=job_id,
        urls=urls,
        schema_dict=schema_dict,
        max_depth=min(max_depth, 5),
        max_pages=min(max_pages, 200),
    )

    return {
        "job_id": job_id,
        "total_urls": len(urls),
        "urls_found": urls[:10],  # preview first 10
        "status": "running",
        "poll_url": f"/api/v1/webcrawl/{job_id}",
    }


@router.get("/{job_id}")
async def get_webcrawl_status(job_id: str):
    """Poll crawl job status."""
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(404, f"Job '{job_id}' not found.")
    return {
        "job_id": job_id,
        "status": job["status"],
        "total": job["total"],
        "completed": job["completed"],
        "failed": job["failed"],
        "result_count": len(job["results"]),
        "schema_name": job["schema_name"],
        "created_at": job["created_at"],
    }


@router.get("/{job_id}/results")
async def get_webcrawl_results(job_id: str):
    """Get full results for a completed crawl job."""
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(404, f"Job '{job_id}' not found.")
    return {
        "job_id": job_id,
        "status": job["status"],
        "results": job["results"],
        "schema_name": job["schema_name"],
    }


@router.get("/{job_id}/excel")
async def export_webcrawl_excel(job_id: str):
    """Download crawl results as Excel."""
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(404, f"Job '{job_id}' not found.")
    if job["status"] == "running":
        raise HTTPException(400, "Job still running.")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed.")

    results = job["results"]
    if not results:
        raise HTTPException(400, "No results to export.")

    # Build columns
    all_keys = ["source_url", "page_title"]
    for r in results:
        for k in r:
            if k not in all_keys and k not in ("error",):
                all_keys.append(k)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Web Crawl Results"

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    row_fill_even = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Calibri", size=10)

    for col_idx, col in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    for row_idx, result in enumerate(results, 2):
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        for col_idx, col in enumerate(all_keys, 1):
            val = result.get(col)
            if val is None:
                val = ""
            elif isinstance(val, (list, dict)):
                val = json.dumps(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin

    for col_idx, col in enumerate(all_keys, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col)), max((len(str(r.get(col, "") or "")) for r in results), default=0))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=webcrawl_{job_id[:8]}.xlsx"},
    )


@router.get("/{job_id}/csv")
async def export_webcrawl_csv(job_id: str):
    """Download crawl results as CSV."""
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(404, f"Job '{job_id}' not found.")

    results = job["results"]
    if not results:
        raise HTTPException(400, "No results to export.")

    all_keys = ["source_url", "page_title"]
    for r in results:
        for k in r:
            if k not in all_keys and k not in ("error",):
                all_keys.append(k)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction="ignore")
    writer.writeheader()
    for r in results:
        writer.writerow({k: "" if r.get(k) is None else r.get(k) for k in all_keys})

    content = output.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=webcrawl_{job_id[:8]}.csv"},
    )


# ── Background processor ──────────────────────────────────────────────────────

async def _process_crawl(
    job_id: str,
    urls: list[str],
    schema_dict: dict,
    max_depth: int,
    max_pages: int,
):
    """Process all URLs in the background."""
    from app.services.web_crawler import crawl_url_list

    job = _jobs.get(job_id)
    if not job:
        return

    try:
        def on_progress(i, total, url):
            if job_id in _jobs:
                _jobs[job_id]["completed"] = i

        results = await crawl_url_list(
            urls=urls,
            schema=schema_dict,
            max_depth=max_depth,
            on_progress=on_progress,
        )

        _jobs[job_id]["results"] = results
        _jobs[job_id]["completed"] = len(urls)
        _jobs[job_id]["status"] = "completed"

    except Exception as e:
        _jobs[job_id]["status"] = "failed"
        _jobs[job_id]["error"] = str(e)


# ── URL extractor from file ───────────────────────────────────────────────────

def _extract_urls_from_file(content: bytes, filename: str) -> list[str]:
    """Extract URLs from CSV, Excel, JSON, or plain text file."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    urls = []

    if ext in ("xlsx", "xls"):
        try:
            import openpyxl
            import io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str) and cell.startswith("http"):
                            urls.append(cell.strip())
        except Exception:
            pass

    elif ext == "json":
        try:
            data = json.loads(content.decode("utf-8", errors="ignore"))
            urls = _extract_urls_from_json(data)
        except Exception:
            pass

    elif ext == "csv":
        try:
            text = content.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                for cell in row:
                    cell = cell.strip()
                    if cell.startswith("http"):
                        urls.append(cell)
        except Exception:
            pass

    else:
        # Plain text — one URL per line
        text = content.decode("utf-8", errors="ignore")
        for line in text.splitlines():
            line = line.strip()
            if line.startswith("http"):
                urls.append(line)

    # Deduplicate and validate
    seen = set()
    clean = []
    for url in urls:
        url = url.strip().rstrip(",;")
        if url not in seen and url.startswith("http"):
            seen.add(url)
            clean.append(url)

    return clean


def _extract_urls_from_json(data) -> list[str]:
    """Recursively extract all URLs from any JSON structure."""
    urls = []
    if isinstance(data, str):
        if data.startswith("http"):
            urls.append(data.strip())
    elif isinstance(data, list):
        for item in data:
            urls.extend(_extract_urls_from_json(item))
    elif isinstance(data, dict):
        for value in data.values():
            urls.extend(_extract_urls_from_json(value))
    return urls
